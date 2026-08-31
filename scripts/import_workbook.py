"""将账套助手导出的 Excel 原始工作表转换为看板可读取的只读 JSON。

这一步只提取单元格值和工作表边界，不改写源 Excel；生产环境由各账套助手调用同一数据契约生成原始资料包。
"""
from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl


SHEET_MAP = {
    "balance_sheet": ["资产负债表"],
    "income_statement": ["利润表"],
    "cash_flow": ["现金流量表-钱去向", "现金流量表"],
    "journal": ["序时账", "7月序时账"],
    "trial_balance": ["科目余额表", "7月科目余额表"],
}


def serialise(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def extract_sheet(workbook, candidates):
    sheet_name = next((name for name in candidates if name in workbook.sheetnames), None)
    if not sheet_name:
        return None
    sheet = workbook[sheet_name]
    max_row = min(sheet.max_row, 5000)
    max_col = min(sheet.max_column, 40)
    rows = []
    for row_index in range(1, max_row + 1):
        values = [serialise(sheet.cell(row_index, col_index).value) for col_index in range(1, max_col + 1)]
        if any(value not in (None, "") for value in values):
            rows.append({"row": row_index, "cells": values})
    return {"sourceSheet": sheet_name, "maxRow": sheet.max_row, "maxCol": sheet.max_column, "rows": rows}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    workbook = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    payload = {key: extract_sheet(workbook, candidates) for key, candidates in SHEET_MAP.items()}
    payload["sourceFile"] = str(args.source)
    payload["sourceType"] = "账套助手原始 Excel"
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({key: bool(value) for key, value in payload.items() if key in SHEET_MAP}, ensure_ascii=False))


if __name__ == "__main__":
    main()
